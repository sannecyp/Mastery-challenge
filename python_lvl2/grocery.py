import Tkinter as tk
from collections import deque
from openpyxl import Workbook
from openpyxl import load_workbook
import os.path
 

class Create_Window(tk.Frame):
	P_add=35
	Prathik_buttons=[]
	Prathik_labels=[]
	Prathik_remove=[]
	
	Chet_buttons=[]
	Chet_labels=[]
	Chet_remove=[]
	
	Ram_buttons=[]
	Ram_labels=[]
	Ram_remove=[]
	
	Dennis_buttons=[]
	Dennis_labels=[]
	Dennis_remove=[]
	Value=[]
	Prathik_row={}

	# open and load excel file, or create one if it doesnt exist
	if os.path.isfile('balances.xlsx'):
		wb1 = load_workbook('balances.xlsx')
	else:
		wb1 =Workbook()
		wb1.save('balances.xlsx')
	def __init__(self,root):
		# Create a frame and an inside canvas
		tk.Frame.__init__(self, root)
		
		# self.hsb= Scrollbar(root, orient="horizontal" )
		# self.hsb.config(command=self.xview)
		# root.configure(xscrollcommand=self.hsb.set)
		# self.hsb.pack(side="bottom" , fill=Y)
		self.root=root
		self.canvas5 = tk.Canvas(root, borderwidth=0, background="#ffffff",height=350,width=1000)
		self.frame5 = tk.Frame(self.canvas5)
		
		self.canvas5.pack(side="bottom", fill="both", expand=0)
		
		# Create four windows/scrollbars (one per person)
		self.canvas5.create_window((4,4), window=self.frame5, anchor="nw", 
								  tags="self.frame5")
		
		
		self.canvas = tk.Canvas(root, borderwidth=0, background="#ffffff",height=200,width=225)
		self.frame = tk.Frame(self.canvas)
		self.vsb = tk.Scrollbar(root, orient="vertical", command=self.canvas.yview)
		self.canvas.configure(yscrollcommand=self.vsb.set)

		self.canvas.pack(side="left", fill="both", expand=1)
		self.vsb.pack(side="left", fill="y")
		self.canvas.create_window((4,4), window=self.frame, anchor="nw", 
								  tags="self.frame")
		
		self.frame.bind("<Configure>", self.onFrameConfigure)
		
		self.canvas2 = tk.Canvas(root, borderwidth=0, background="#ffffff",height=200,width=225)
		self.frame2 = tk.Frame(self.canvas2)
		self.vsb2 = tk.Scrollbar(root, orient="vertical", command=self.canvas2.yview)
		self.canvas2.configure(yscrollcommand=self.vsb2.set)
		
		self.canvas2.pack(side="left", fill="both", expand=1)
		self.vsb2.pack(side="left", fill="y")
		
		self.canvas2.create_window((4,4), window=self.frame2, anchor="nw", 
								  tags="self.frame2")

		self.frame2.bind("<Configure>", self.onFrameConfigure)
		
		
		self.canvas3 = tk.Canvas(root, borderwidth=0, background="#ffffff",height=200,width=225)
		self.frame3 = tk.Frame(self.canvas3)
		self.vsb3 = tk.Scrollbar(root, orient="vertical", command=self.canvas3.yview)
		self.canvas3.configure(yscrollcommand=self.vsb3.set)
		
		self.canvas3.pack(side="left", fill="both", expand=1)
		self.vsb3.pack(side="left", fill="y")
		
		self.canvas3.create_window((4,4), window=self.frame3, anchor="nw", 
								  tags="self.frame3")

		self.frame3.bind("<Configure>", self.onFrameConfigure)
		
		
		self.canvas4 = tk.Canvas(root, borderwidth=0, background="#ffffff",height=200,width=225)
		self.frame4 = tk.Frame(self.canvas4)
		self.vsb4 = tk.Scrollbar(root, orient="vertical", command=self.canvas4.yview)
		self.canvas4.configure(yscrollcommand=self.vsb4.set)
		
		self.canvas4.pack(side="left", fill="both", expand=1)
		self.vsb4.pack(side="left", fill="y")
		
		self.canvas4.create_window((4,4), window=self.frame4, anchor="nw", 
								  tags="self.frame4")

		self.frame4.bind("<Configure>", self.onFrameConfigure)
		
		

		self.initUI()
	
	# Set all frames to be scrollable
	def onFrameConfigure(self, event):
		
		self.canvas.configure(scrollregion=self.canvas.bbox("all"))
		self.canvas2.configure(scrollregion=self.canvas2.bbox("all"))
		self.canvas3.configure(scrollregion=self.canvas3.bbox("all"))
		self.canvas4.configure(scrollregion=self.canvas4.bbox("all"))
		
	# Save the balances in the excel worksheet, one person per column
	def save_ws(self):
		sheet_ranges=self.wb1['Sheet']
		
		for x in range(1,200): 
			sheet_ranges[('A'+str(x))]=None
			sheet_ranges[('C'+str(x))]=None
			sheet_ranges[('E'+str(x))]=None
			sheet_ranges[('G'+str(x))]=None
			
		i=1
		for x in self.Prathik_labels:
			cell='A'+str(i)
			value=x.get()
			try:
				sheet_ranges[cell]=float(value)
			except:
				pass
			sheet_ranges[cell].number_format
			i+=1
		
		i=1
		for x in self.Chet_labels:
			cell='C'+str(i)
			value=x.get()
			try:
				sheet_ranges[cell]=float(value)
			except:
				pass
			sheet_ranges[cell].number_format
			i+=1
		
		i=1
		for x in self.Ram_labels:
			cell='E'+str(i)
			value=x.get()
			try:
				sheet_ranges[cell]=float(value)
			except:
				pass
			sheet_ranges[cell].number_format
			i+=1

		i=1
		for x in self.Dennis_labels:
			cell='G'+str(i)
			value=x.get()
			try:
				sheet_ranges[cell]=float(value)
			except:
				pass
			sheet_ranges[cell].number_format
			i+=1
		
		self.wb1.save('balances.xlsx')
	
	# Close the application
	def quit(self):
		 self.root.destroy()
	
	# Create labels for names, save and quit button
	def initUI(self):
		lent=len(self.Prathik_buttons)
		y_offset=35
		
		# self.root.title("Groceries")
		# self.pack(fill=BOTH, expand=1)
		
		
		self.frame.grid_rowconfigure(0, weight=1)
		self.frame.grid_columnconfigure(0, weight=1)
		self.frame.grid_columnconfigure(1, weight=1)
		self.frame.grid_columnconfigure(2, weight=1)
		
		self.frame2.grid_rowconfigure(0, weight=1)
		self.frame2.grid_columnconfigure(0, weight=1)
		self.frame2.grid_columnconfigure(1, weight=1)
		self.frame2.grid_columnconfigure(2, weight=1)
		
		self.frame3.grid_rowconfigure(0, weight=1)
		self.frame3.grid_columnconfigure(0, weight=1)
		self.frame3.grid_columnconfigure(1, weight=1)
		self.frame3.grid_columnconfigure(2, weight=1)
		
		self.frame4.grid_rowconfigure(0, weight=1)
		self.frame4.grid_columnconfigure(0, weight=1)
		self.frame4.grid_columnconfigure(1, weight=1)
		self.frame4.grid_columnconfigure(2, weight=1)
		
		quitButton = tk.Button(self.canvas5, text="Quit",command=self.quit)
		# quitButton.place(x=1050, y=400)
		quitButton.pack(side="right", fill="x", expand=0)
		
		saveButton = tk.Button(self.canvas5, text="Save",command=self.save_ws)
		saveButton.pack(side="right", fill="x", expand=0)
		
		
		
		# self.pack()
		
		Calc_button = tk.Button(self.canvas5, text="Calculate",command=lambda: self.calculate())
		Calc_button.pack(side="left", fill="x", expand=0)
		
		Prathik = tk.Label(self.frame, text="Prathik")
		# Prathik.pack(side="top",fill="y")
		Prathik.grid(row=0, column=0, columnspan=3, sticky="we")
		
		Chet = tk.Label(self.frame2, text="Chethiya")
		Chet.grid(row=0, column=0, columnspan=3, sticky="we")
		
		Ram = tk.Label(self.frame3, text="Ram")
		Ram.grid(row=0, column=0, columnspan=3, sticky="we")
		
		Ram = tk.Label(self.frame4, text="Dennis")
		Ram.grid(row=0, column=0, columnspan=3, sticky="we")
		# self.grid_columnconfigure(3, weight=1)
		# self.Prathik_buttons.append(Button(self, text="add",command=self.addbutton(len(self.Prathik_buttons))))
		
		
		
		self.Prathik_buttons.append(tk.Button(self.frame, text="add",command=lambda Prathik_lent=len(self.Prathik_buttons): self.addbutton(Prathik_lent,0, self.Prathik_buttons,self.Prathik_labels,self.Prathik_remove)))
		self.Prathik_buttons[0].grid(row=3, column=0)
		self.Prathik_labels.append(tk.Entry(self.frame))
		self.Prathik_labels[0].grid(row=3, column=1,sticky="we")
		self.Prathik_remove.append(tk.Button(self.frame, text="remove",command=lambda Prathik_lent=len(self.Prathik_buttons): self.removebutton(Prathik_lent,0, self.Prathik_buttons,self.Prathik_labels,self.Prathik_remove)))
		self.Prathik_remove[0].grid(row=3, column=2, columnspan=1)
		
		self.Chet_buttons.append(tk.Button(self.frame2, text="add",command=lambda Chet_lent=len(self.Chet_buttons): self.addbutton(Chet_lent,1, self.Chet_buttons,self.Chet_labels,self.Chet_remove)))
		self.Chet_buttons[0].grid(row=3, column=0)
		self.Chet_labels.append(tk.Entry(self.frame2))
		self.Chet_labels[0].grid(row=3, column=1,sticky="we")
		self.Chet_remove.append(tk.Button(self.frame2, text="remove",command=lambda Chet_lent=len(self.Chet_buttons): self.removebutton(Chet_lent,1, self.Chet_buttons,self.Chet_labels,self.Chet_remove)))
		self.Chet_remove[0].grid(row=3, column=2, columnspan=1)
		
		self.Ram_buttons.append(tk.Button(self.frame3, text="add",command=lambda Ram_lent=len(self.Ram_buttons): self.addbutton(Ram_lent,2, self.Ram_buttons,self.Ram_labels,self.Ram_remove)))
		self.Ram_buttons[0].grid(row=3, column=0)
		self.Ram_labels.append(tk.Entry(self.frame3))
		self.Ram_labels[0].grid(row=3, column=1,sticky="we")
		self.Ram_remove.append(tk.Button(self.frame3, text="remove",command=lambda Ram_lent=len(self.Ram_buttons): self.removebutton(Ram_lent,2, self.Ram_buttons,self.Ram_labels,self.Ram_remove)))
		self.Ram_remove[0].grid(row=3, column=2, columnspan=1)
		
		self.Dennis_buttons.append(tk.Button(self.frame4, text="add",command=lambda Dennis_lent=len(self.Dennis_buttons): self.addbutton(Dennis_lent,3, self.Dennis_buttons,self.Dennis_labels,self.Dennis_remove)))
		self.Dennis_buttons[0].grid(row=3, column=0)
		self.Dennis_labels.append(tk.Entry(self.frame4))
		self.Dennis_labels[0].grid(row=3, column=1,sticky="we")
		self.Dennis_remove.append(tk.Button(self.frame4, text="remove",command=lambda Dennis_lent=len(self.Dennis_buttons): self.removebutton(Dennis_lent,3, self.Dennis_buttons,self.Dennis_labels,self.Dennis_remove)))
		self.Dennis_remove[0].grid(row=3, column=2, columnspan=1)
		
		
		self.Value.append(self.Prathik_remove[0])
		self.Prathik_row[0]=self.Value[0]
		
		self.importdata()
		
		self.calculate()
		
	# Import the spreadsheet into the gui. Each value has an add and remove button attached
	def importdata(self):
		sheet_ranges=self.wb1['Sheet']
		Prathik_values=[]
		for row in sheet_ranges.iter_rows('A1:A200'): 
			if (row[0].value != None) and (row[0].value !=0):
				Prathik_values.append(row[0].value)
		i=0
		for row in Prathik_values:
			self.addbutton(i,0, self.Prathik_buttons,self.Prathik_labels,self.Prathik_remove)
			label_row=self.Prathik_labels[i]
			label_row.insert( "end",str(row))
			i+=1
		
		Chet_values=[]
		for row in sheet_ranges.iter_rows('C1:C200'): 
			if (row[0].value != None) and (row[0].value !=0):
				Chet_values.append(row[0].value)
		i=0
		for row in Chet_values:
			self.addbutton(i,1, self.Chet_buttons,self.Chet_labels,self.Chet_remove)
			label_row=self.Chet_labels[i]
			label_row.insert( "end",str(row))
			i+=1
			
		Ram_values=[]
		for row in sheet_ranges.iter_rows('E1:E200'): 
			if (row[0].value != None) and (row[0].value !=0):
				Ram_values.append(row[0].value)
		i=0
		for row in Ram_values:
			self.addbutton(i,2, self.Ram_buttons,self.Ram_labels,self.Ram_remove)
			label_row=self.Ram_labels[i]
			label_row.insert( "end",str(row))
			i+=1
			
		Dennis_values=[]
		for row in sheet_ranges.iter_rows('G1:G200'): 
			if (row[0].value != None) and (row[0].value !=0):
				Dennis_values.append(row[0].value)
		i=0
		for row in Dennis_values:
			self.addbutton(i,3, self.Dennis_buttons,self.Dennis_labels,self.Dennis_remove)
			label_row=self.Dennis_labels[i]
			label_row.insert( "end",str(row))
			i+=1
		
	# Add a button into an arbitrarry place in the gui
	def addbutton_index_change(self,index,lent,array):
		if ((lent)-(index+1))>=0:	
			for row in reversed(array[(index+1):(lent)]):
				array[array.index(row)+1],array[array.index(row)]=array[array.index(row)],array[array.index(row)+1]
			
	# Rearrange all the values if a value is removed
	def reposition(self,index,horizontal,array):
		for cell in array:
			# self.canvas.grid_rowconfigure((array.index(cell)+1), weight=1)
			cell.grid(row=(array.index(cell)+3),column=horizontal,sticky="we")
		
		
			# for cell in array:
				# if cell.winfo_height()<20:
					# cell.config( height = 20)
		
	# Add a value to an arbitrary person, and adjust the buttons accordingly
	def addbutton(self, index,hrz, buttons,labels,remove):
		lent=len(buttons)
		
		if hrz==0:
			# print index
			# print lent
			
			# self.canvas.grid_rowconfigure(lent+1, weight=1)
			buttons.append(tk.Button(self.frame, text="add",command= lambda lent=len(buttons): self.addbutton(lent,hrz, buttons,labels,remove )))
			buttons[-1].grid(row=lent+3,column=0,sticky="we")
			labels.append(tk.Entry(self.frame))
			labels[-1].grid(row=lent+3,column=1,sticky="we")
			remove.append(tk.Button(self.frame, text="remove",command=lambda lent=len(buttons): self.removebutton(lent,hrz, buttons,labels,remove)))
			remove[-1].grid(row=lent+3,column=2,sticky="we")
			
		if hrz==1:
			# print index
			# print lent
			
			# self.canvas.grid_rowconfigure(lent+1, weight=1)
			buttons.append(tk.Button(self.frame2, text="add",command= lambda lent=len(buttons): self.addbutton(lent,hrz, buttons,labels,remove )))
			buttons[-1].grid(row=lent+3,column=0,sticky="we")
			labels.append(tk.Entry(self.frame2))
			labels[-1].grid(row=lent+3,column=1,sticky="we")
			remove.append(tk.Button(self.frame2, text="remove",command=lambda lent=len(buttons): self.removebutton(lent,hrz, buttons,labels,remove)))
			remove[-1].grid(row=lent+3,column=2,sticky="we")
			
		if hrz==2:
			# print index
			# print lent
			
			# self.canvas.grid_rowconfigure(lent+1, weight=1)
			buttons.append(tk.Button(self.frame3, text="add",command= lambda lent=len(buttons): self.addbutton(lent,hrz, buttons,labels,remove )))
			buttons[-1].grid(row=lent+3,column=0,sticky="we")
			labels.append(tk.Entry(self.frame3))
			labels[-1].grid(row=lent+3,column=1,sticky="we")
			remove.append(tk.Button(self.frame3, text="remove",command=lambda lent=len(buttons): self.removebutton(lent,hrz, buttons,labels,remove)))
			remove[-1].grid(row=lent+3,column=2,sticky="we")
			
		if hrz==3:
			# print index

			# print lent
			
			# self.canvas.grid_rowconfigure(lent+1, weight=1)
			buttons.append(tk.Button(self.frame4, text="add",command= lambda lent=len(buttons): self.addbutton(lent,hrz, buttons,labels,remove )))
			buttons[-1].grid(row=lent+3,column=0,sticky="we")
			labels.append(tk.Entry(self.frame4))
			labels[-1].grid(row=lent+3,column=1,sticky="we")
			remove.append(tk.Button(self.frame4, text="remove",command=lambda lent=len(buttons): self.removebutton(lent,hrz, buttons,labels,remove)))
			remove[-1].grid(row=lent+3,column=2,sticky="we")
			
		self.addbutton_index_change(index,lent,labels)
		self.reposition(index,0,buttons)
		self.reposition(index,1,labels)
		self.reposition(index,2,remove)

	# Remove an arbitrary value from an arbitrary person, and rearrange the buttons accordingly
	def removebutton(self, index,hrz, buttons,labels,remove):
		# print "index %i" %index
		
		lent=len(buttons)
		
		if lent>1:
			buttons[-1].destroy()
			remove[-1].destroy()
			
			buttons.pop()
			remove.pop()
			
			labels[index-1].destroy()
			labels.pop(index-1)
		
		
			self.reposition(index,0,buttons)
			self.reposition(index,1,labels)
			self.reposition(index,2,remove)
		
	# Do the math to find out how much each person owes
	def calculate(self):
		sheet_ranges=self.wb1['Sheet']
		

		Prathik_value=[]
		Chet_value=[]
		Ram_value=[]
		Dennis_value=[]

		for x in self.Prathik_labels:
			value=x.get()
			try:
				Prathik_value.append(float(value))
			except:
				pass

		for x in self.Chet_labels:
			value=x.get()
			try:
				Chet_value.append(float(value))
			except:
				pass
		
		for x in self.Ram_labels:
			value=x.get()
			try:
				Ram_value.append(float(value))
			except:
				pass
				
		for x in self.Dennis_labels:
			value=x.get()
			try:
				Dennis_value.append(float(value))
			except:
				pass
		
		self.Prathik_sum=sum(Prathik_value)
		self.Chet_sum=sum(Chet_value)
		self.Ram_sum=sum(Ram_value)
		self.Dennis_sum=sum(Dennis_value)
		
		self.Average=(self.Prathik_sum+self.Chet_sum+self.Ram_sum+self.Dennis_sum)/4
		
		# Display the value on the gui
		Prathik_paid = tk.Label(self.frame, text=("Prathik paid: "+str(self.Prathik_sum)))
		Prathik_paid.grid(row=1, column=0, columnspan=3, sticky="we")
		Prathik_owes = tk.Label(self.frame, text=("Prathik owes: "+str(-self.Prathik_sum+self.Average)))
		Prathik_owes.grid(row=2, column=0, columnspan=3, sticky="we")
		
		Chet_paid = tk.Label(self.frame2, text=("Chet paid: "+str(self.Chet_sum)))
		Chet_paid.grid(row=1, column=0, columnspan=3, sticky="we")
		Chet_owes = tk.Label(self.frame2, text=("Chet owes: "+str(-self.Chet_sum+self.Average)))
		Chet_owes.grid(row=2, column=0, columnspan=3, sticky="we")
		
		Ram_paid = tk.Label(self.frame3, text=("Ram paid: "+str(self.Ram_sum)))
		Ram_paid.grid(row=1, column=0, columnspan=3, sticky="we")
		Ram_owes = tk.Label(self.frame3, text=("Ram owes: "+str(-self.Ram_sum+self.Average)))
		Ram_owes.grid(row=2, column=0, columnspan=3, sticky="we")

		Dennis_paid = tk.Label(self.frame4, text=("Dennis paid: "+str(self.Dennis_sum)))
		Dennis_paid.grid(row=1, column=0, columnspan=3, sticky="we")
		Dennis_owes = tk.Label(self.frame4, text=("Dennis owes: "+str(-self.Dennis_sum+self.Average)))
		Dennis_owes.grid(row=2, column=0, columnspan=3, sticky="we")
		
		
# Create a window, start the gui
def main():
	root=tk.Tk()
	app = Create_Window(root)
	app.mainloop()


if __name__ == '__main__':
	main()
